''' Bereken de cijfers op basis van de normering
    Vul de normering en het aantal punten in bij elke naam.
    Created on 13 nov. 2019
    @author: Tommy Lohn '''

import pandas as pd
import math
import numpy as np
import PySimpleGUI as sg
import os
import os.path
from matplotlib.backends.backend_tkagg import FigureCanvasAgg
import matplotlib.figure
import matplotlib.pyplot as plt
import io
import PIL
import base64
from openpyxl import load_workbook
import matplotlib.patches as mpatches
import numpy.ma as ma
import statistics

import itertools


matplotlib.use('TkAgg')

TESTS = ['T01', 'T02', 'T03', 'T04', 'HT01', 'HT02', 'HT03', 'HT04']
SUFF_GRADE = [5, 5.5, 6]

def print_average_points(points):
    counter = 0
    totalPoints = 0

    for i in points:
        if not math.isnan(i):
            totalPoints = totalPoints + i
            counter += 1
        else:
            totalPoints = totalPoints + 0
            counter += 0

    averagePoints = totalPoints / counter

    averagePoints = round(averagePoints, 2)
    return(averagePoints, counter)

def print_sufficient_percentage(points, norm, maxpoints):

    counter = 0
    totalPoints = 0
    normnumber = float(maxpoints) * (float(norm) / 100)
    suffnumber = 0
    suffpercentage = 0

    for i in points:
        if not math.isnan(i):
            totalPoints = totalPoints + i
            counter += 1
            if i >= normnumber:
                suffnumber += 1
        else:
            totalPoints = totalPoints + 0
            counter += 0

    suffpercentage = suffnumber / counter * 100

    suffpercentage = round(suffpercentage, 2)
    normnumber = round(normnumber, 2)
    suffnumber = round(suffnumber, 2)


    return(suffpercentage, normnumber, suffnumber)

def frange(start, stop, step):
    i = start
    while i < stop:
        yield i
        i = round(i + step, 1)

def draw_plot(norm, maxpoints, instances, normpoints):

    instances.sort()
    plt.hist(instances)
    plt.axvline(x=normpoints, color='r', linestyle='-')
    plt.show(block=False)

def convert_to_bytes(file_or_bytes, resize=None):
    '''
    Will convert into bytes and optionally resize an image that is a file or a base64 bytes object.
    Turns into  PNG format in the process so that can be displayed by tkinter
    :param file_or_bytes: either a string filename or a bytes base64 image object
    :type file_or_bytes:  (Union[str, bytes])
    :param resize:  optional new size
    :type resize: (Tuple[int, int] or None)
    :return: (bytes) a byte-string object
    :rtype: (bytes)
    '''

    plt.close('all')

    if isinstance(file_or_bytes, str):
        img = PIL.Image.open(file_or_bytes)
    else:
        try:
            img = PIL.Image.open(io.BytesIO(base64.b64decode(file_or_bytes)))
        except Exception as e:
            dataBytesIO = io.BytesIO(file_or_bytes)
            img = PIL.Image.open(dataBytesIO)

    cur_width, cur_height = img.size
    if resize:
        new_width, new_height = resize
        scale = min(new_height/cur_height, new_width/cur_width)
        img = img.resize((int(cur_width*scale), int(cur_height*scale)), PIL.Image.ANTIALIAS)
    bio = io.BytesIO()
    img.save(bio, format="PNG")
    del img
    return bio.getvalue()




def save_figure(figure1, figure2, group, test ,normway, norm, path):

    name1 = (test + "-" + group + "-" + normway + "-" + str(norm) + '-normeringslijn' + '.png')
    figure1.savefig(path + '/' + name1)
    name2 = (test + "-" + group + "-" + normway + "-" + str(norm) + '-resultaten' + '.png')
    figure2.savefig(path + '/' + name2)
    return(name1, name2)



def draw_figure(element, figure):
    """
    Draws the previously created "figure" in the supplied Image Element
    :param element: an Image Element
    :param figure: a Matplotlib figure
    :return: The figure canvas
    """

    plt.close('all')  # erases previously drawn plots
    canv = FigureCanvasAgg(figure)
    buf = io.BytesIO()
    canv.print_figure(buf, format='png')
    if buf is None:
        return None
    buf.seek(0)
    element.update(data=buf.read())
    return canv


def draw_normLines(norm, maxpoints, instances, sufficientgrade, normpoints):

    y = []
    x = []
    nanlist = []

    for i in frange(0, 100, 0.1):

        if i <= norm:

            delta_y = sufficientgrade - 1
            delta_x = norm
            a1 = delta_y / delta_x

            y.append(a1 * i + 1)
            x.append(i)

        elif i > norm:

            delta_y = 10 - sufficientgrade
            delta_x = 100 - norm
            a2 = delta_y / delta_x

            b = sufficientgrade - (a2 * norm)

            y.append(a2 * i + b)
            x.append(i)

    sufficientLine = sufficientgrade

    normplot = plt.figure(1)
    plt.title("Normeringslijn")
    plt.xlabel("Percentage goed")
    plt.ylabel("Cijfer")
    plt.axis((0, 100, 0, 10))
    plt.yticks(np.arange(0, 10 + 1, 1.0))
    plt.xticks(np.arange(0, 100 + 1, 10.0))

    plt.axhline(y=sufficientLine, color='b', linestyle=':')
    red_line = mpatches.Patch(color="blue", label="Cijfer voor een voldoende")
    plt.axvline(x=norm, color='r', linestyle=':')
    green_line = mpatches.Patch(color="red", label="Percentage goed voor een voldoende")
    plt.legend(handles=[red_line, green_line])

    mask = ma.masked_less(y, sufficientgrade)
    plt.plot(x, y, "orange")
    plt.plot(x, mask, "green")

    for elem in instances:
        nanlist.append(elem)

    for idx, elem in enumerate(nanlist):
        if math.isnan(elem):
            nanlist.remove(elem)

    scoreplot = plt.figure(2)
    plt.title("Resultaten")
    plt.xlabel("Punten")
    plt.ylabel("Frequentie")
    binvals, bin, patches = plt.hist(nanlist, bins=range(int(min(nanlist) - 1), int(max(nanlist) + 1)), alpha=0.5, ec="black", color="blue")
    plt.axvline(x=normpoints, color='r', linestyle='-')
    red_line2 = mpatches.Patch(color="red", label="Punten goed voor een voldoende")
    plt.legend(handles=[red_line2])

    bin_centers = 0.5 * (bin[:-1] + bin[1:])

    for i, x in zip(patches, bin_centers):
        if float(normpoints) < x < float(maxpoints):
            i.set_facecolor("green")
        else:
            i.set_facecolor("orange")
    #plt.show(block=False)

    return(normplot, scoreplot)

def round_to_half(number):
    return round(number * 2) / 2

def calculate_grades(norm, maxpoints, instances, sufficientgrade, normpoints):

    #y is list with grades, x is list with percentage of points, pointlist is list with points that can be obtained, grades are final grades
    y = []
    x = []
    gradelist = []
    grades = []

    a1 = 0.0
    a2 = 0.0
    b = 0.0

    for j in instances:
        if math.isnan(j):
            grades.append("inh")


        else:
            for i in frange(0, 100+0.1, 0.1):

                if i <= norm:

                    delta_y = sufficientgrade - 1
                    delta_x = norm
                    a1 = delta_y / delta_x

                    y.append(a1 * i + 1)
                    x.append(i)


                elif i > norm:

                    delta_y = 10 - sufficientgrade
                    delta_x = 100 - norm
                    a2 = delta_y / delta_x

                    b = sufficientgrade - (a2 * norm)

                    y.append(a2 * i + b)
                    x.append(i)

            if ((j * 100 / float(maxpoints)) <= norm):
                grades.append(round((float(a1) * (j * 100 / float(maxpoints)) + 1), 1))

            if ((j * 100 / float(maxpoints)) > norm):
                grades.append(round((float(a2) * (j * 100 / float(maxpoints)) + b), 1))

    for i in grades:
        if not i == "inh":
            gradelist.append(i)

    mean_grades = statistics.mean(gradelist)

   # for i in x:
     #   factor = ((i / 100))
     #   pointlist.append(round(factor*float(maxpoints),1))

    #for i in range((int(maxpoints) * 10)+1):
        #i = i/10
       # pointlist.append(i)

   # for i in pointlist:
      #  round_to_half(float(i))


    #for i in instances:
       # if math.isnan(i):
        #    grades.append("inh")

       # else:
          #  grade = y[pointlist.index(i)]
          #  roundedGrade = round(grade, 2)
           # grades.append(round(roundedGrade,1))


    return(grades, mean_grades)

def make_grade_lists(grades, names):
    gradelist=[]
    for idx, val in enumerate(grades):
        gradelist.append([names[idx], val])

    return(gradelist)

def set_column_size(element, size):
    # Only work for sg.Column when `scrollable=True` or `size not (None, None)`
    options = {'width':size[0], 'height':size[1]}
    if element.Scrollable or element.Size!=(None, None):
        element.Widget.canvas.configure(**options)
    else:
        element.Widget.pack_propagate(0)
        element.set_size(size)


def add_column(wb, sheet_name, column):
    ws = wb[sheet_name]
    new_column = 3

    for rowy, value in enumerate(column, start=1):
        ws.cell(row=rowy, column=new_column, value=value)

def runApp():

    sg.theme("darkteal7")

    file_list_column = [
        [sg.Text('Vul de volgende variabelen in')],
        [sg.Text('Type normering', size=(20, 1)), sg.Radio('percentage', "RADIOB", default=True, key='-RADIOB1-'), sg.Radio('punten', "RADIOB", default=False, key='RADIOB2')],
        [sg.Text('Normeringspercentage', size=(20, 1)), sg.Slider(range=(1, 100), orientation='h', size=(35.5, 20), default_value=70)],
        [sg.Text('Normeringspunten', size=(20, 1)), sg.InputText()],
        [sg.Text('Maximaal aantal punten', size=(20, 1)), sg.InputText()],
        [sg.Text('Welk cijfer is een voldoende', size=(20, 1)), sg.InputOptionMenu((SUFF_GRADE))],
        [sg.Text('Toets nummer', size=(20, 1)), sg.InputOptionMenu((TESTS))],
        [sg.Text('Kies hier het excel bestand')], [sg.In(size=(25, 1), enable_events=True, key="-FOLDER-"), sg.FolderBrowse("Bladeren")],
    [
            sg.Listbox(
                values=[], enable_events=True, size=(40, 20), key="-FILE LIST-"
            )
        ],
        [sg.Submit("Resultaat", key="-SUBMIT-"), sg.Button("Voer cijfers in", key="-GRADEBUTTON-"), sg.Button("Sla grafieken op", key="-SAVEPIC-")],
    ]

    # For now will only show the name of the file that was chosen
    Results = [
        [sg.Text(size=(35, 1), key="-LTEXT-")],
        [sg.Text(size=(35, 1), key="-NUMP-"), sg.Text(size=(15, 1), key="-NUMP2-")],
        [sg.Text(size=(35, 1), key="-AVG-"), sg.Text(size=(15, 1), key="-AVG2-")],
        [sg.Text(size=(35, 1), key="-SUFP-"), sg.Text(size=(15, 1), key="-SUFP2-")],
        [sg.Text(size=(35, 1), key="-SNUM-"), sg.Text(size=(15, 1), key="-SNUM2-")],
        [sg.Text(size=(35, 1), key="-SUF%-"), sg.Text(size=(15, 1), key="-SUF%2-")],
        [sg.Text(size=(35, 1), key="-MEANT-"), sg.Text(size=(15, 1), key="-MEAN-")],
        # [sg.Text(size=(35, 1), key="-GRADES-")],
        # [sg.Text(size=(35, 6), key="-GRADES2-")],
        [sg.Frame(layout=[
            [sg.Table(values=[['', '']], headings=["Namen", "Cijfers"], num_rows=(10), row_height=17,
                      auto_size_columns=False, hide_vertical_scroll=False, key='-TABLE-')]],
            key="-FRAME-", title='CijferLijst', title_color='white', relief=sg.RELIEF_SUNKEN, tooltip='Frame',
            visible=False)],
        [sg.Text("Uitleg van de app:\n\n"
                 "Met deze app kan je de ideale normering kiezen per toets.\n"
                 "De app berekent verschillende handige waardes\n"
                 "en er is de optie om alle cijfers weg te schrijven\n"
                 "\n"
                 "Er zijn twee manieren om de normering te bepalen:\n"
                 "-1- percentage van de punten dat een leerling goed moet hebben \n"
                 " voor een voldoende\n"
                 "-2- punten die een leerling moet halen voor een voldoende\n"
                 "Het type normering wordt bepaald door de keuzeknoppen bovenaan\n"
                 "\n"
                 "Verder moet het maximale aantal punten worden ingevoerd\n"
                 "\n"
                 "Kies een cijfer dat een voldoende is (5, 5.5, 6)"
                 "\n"
                 "Kies een toets (T01, T02, T03, T04)\n"
                 "\n"
                 "U kiest een map waar uw leerjaar in staat \n"
                 "door middel van de bladeren knop\n"
                 "\n"
                 "Alle klassen uit dit leerjaar zullen in de lijst verschiijnen\n"
                 "Kies uit deze lijst een klas en druk vervolgens op resultaat\n"
                 "\n"
                 "U kunt de cijfers invoeren door op de "'"voer cijfers in knop"'" te drukken\n"
                 "U kunt de grafieken opslaan door op de "'"sla grafieken op knop"'" te drukken"
                 , key="-RTEXT-")],

    ]

    Images = [
        [sg.Image(key='-IMAGE1-', size=(20,10))],
        [sg.Image(key='-IMAGE2-', size=(20,10))],
    ]
    # ----- Full layout -----
    layout = [
        [
            sg.Column(file_list_column, key='-LEFTCOLUMN-'),
            sg.VSeperator(),
            sg.Column(Results, key='-CENTERCOLUMN-'),
            sg.VSeperator(),
            sg.Column(Images, key='-RIGHTCOLUMN-'),
        ]
    ]

    window = sg.Window("Nakijken", layout, finalize=True, resizable=True)
    window.Maximize()
    window_width, window_height = window.size
    set_column_size(window['-LEFTCOLUMN-'], (window_width / 3, window_height))
    set_column_size(window['-CENTERCOLUMN-'], (window_width / 3.5, window_height))
    set_column_size(window['-RIGHTCOLUMN-'], (window_width / 3, window_height))

    #window['-FILE LIST-'].expand(True, True)


    image_element = window['-IMAGE1-']  # type: sg.Image
    image_element2 = window['-IMAGE2-'] # type: sg.Image

    # Run the Event Loop
    while True:
        event, values = window.read()
        if event == "Exit" or event == sg.WIN_CLOSED:
            break
        # Folder name was filled in, make a list of files in the folder
        if event == "-FOLDER-":
            folder = values["-FOLDER-"]
            try:
                # Get list of files in folder
                file_list = os.listdir(folder)
            except:
                file_list = []

            fnames = [
                f
                for f in file_list
                if os.path.isfile(os.path.join(folder, f))
                and f.lower().endswith(".xlsx")
            ]
            window["-FILE LIST-"].update(fnames)
        elif event == "-FILE LIST-":  # A file was chosen from the listbox
            try:
                filename = os.path.join(
                    values["-FOLDER-"], values["-FILE LIST-"][0]
                )
                wb = load_workbook(filename)
                data = pd.ExcelFile(filename)
                sheet1 = pd.read_excel(data, 'T01')
                sheet2 = pd.read_excel(data, 'T02')
                sheet3 = pd.read_excel(data, 'T03')
                sheet4 = pd.read_excel(data, 'T04')
                sheet5 = pd.read_excel(data, 'HT01')
                sheet6 = pd.read_excel(data, 'HT02')
                sheet7 = pd.read_excel(data, 'HT03')
                sheet8 = pd.read_excel(data, 'HT04')
            except:
                pass

        elif event == "-SUBMIT-":
            try:

                window_width, window_height = window.size

                if values[4] == '':
                    sg.popup("Kies een toetsnummer")

                if values[4] == "T01":
                    POINTS = sheet1['Punten'].tolist()
                    NAMES = sheet1['Naam'].tolist()
                if values[4] == "T02":
                    POINTS = sheet2['Punten'].tolist()
                    NAMES = sheet2['Naam'].tolist()
                if values[4] == "T03":
                    POINTS = sheet3['Punten'].tolist()
                    NAMES = sheet3['Naam'].tolist()
                if values[4] == "T04":
                    POINTS = sheet4['Punten'].tolist()
                    NAMES = sheet4['Naam'].tolist()

                if values[4] == "HT01":
                    POINTS = sheet5['Punten'].tolist()
                    NAMES = sheet5['Naam'].tolist()
                if values[4] == "HT02":
                    POINTS = sheet6['Punten'].tolist()
                    NAMES = sheet6['Naam'].tolist()
                if values[4] == "HT03":
                    POINTS = sheet7['Punten'].tolist()
                    NAMES = sheet7['Naam'].tolist()
                if values[4] == "HT04":
                    POINTS = sheet8['Punten'].tolist()
                    NAMES = sheet8['Naam'].tolist()

                if values[2] == '':
                    sg.popup("vul maximaal aantal punten in")

                if values[1] == '' and values['RADIOB2'] == True:
                    sg.popup("vul normeringspunten in")

                if values[3] == '':
                    sg.popup("kies een voldoende cijfer in")

                if values['RADIOB2'] == True:
                    if float(values[1]) > float(values[2]):
                        sg.popup("normeringspunten kunnen niet hoger dan het maximaal aantal punten zijn")

                for i in POINTS:
                    if float(i) > float(values[2]):
                        sg.popup("er is een cijfer hoger dan het maximaal aantal punten")

                print(POINTS)


                if values['RADIOB2'] == True:
                    values[0] = (float(values[1]) / float(values[2]) * 100)
                    normway = "punten"
                else:
                    normway = "percentage"

                fig1 = draw_normLines(values[0], values[2], POINTS, float(values[3]),
                                                                       print_sufficient_percentage(POINTS, values[0],
                                                                                                    values[2])[1])[0]
                fig2 = draw_normLines(values[0], values[2], POINTS, float(values[3]),
                                                                       print_sufficient_percentage(POINTS, values[0],
                                                                                                    values[2])[1])[1]

                excelfile = str(filename).split("\\")[-1]
                group = str(excelfile).split(".")[0]

                window["-LTEXT-"].update("Hieronder de resultaten")
                window["-RTEXT-"].update("")

                window["-NUMP-"].update("Aantal mensen dat de toets hebben gemaakt:")
                window["-NUMP2-"].update(print_average_points(POINTS)[1])

                window["-AVG-"].update("gemiddeld aantal punten:")
                window["-AVG2-"].update(print_average_points(POINTS)[0])

                window["-SUFP-"].update("Aantal punten voor een voldoende:")
                window["-SUFP2-"].update(print_sufficient_percentage(POINTS, values[0], values[2])[1])

                window["-SNUM-"].update("Aantal mensen met een voldoende:")
                window["-SNUM2-"].update(print_sufficient_percentage(POINTS, values[0], values[2])[2])

                window["-SUF%-"].update("Percentage met een voldoende:")
                window["-SUF%2-"].update(print_sufficient_percentage(POINTS, values[0], values[2])[0])

                window["-MEANT-"].update("gemiddeld cijfer:")
                window["-MEAN-"].update(calculate_grades(values[0], values[2], POINTS, float(values[3]), print_sufficient_percentage(POINTS, values[0], values[2])[1])[1])

                #window["-GRADES-"].update("cijfers:")
                #window["-GRADES2-"].update(calculate_grades(values[0], values[2], POINTS, float(values[3]), print_sufficient_percentage(POINTS, values[0], values[2])[1]))

                window["-FRAME-"].update(visible=True)
                window['-TABLE-'].update(values=make_grade_lists(calculate_grades(values[0], values[2], POINTS, float(values[3]), print_sufficient_percentage(POINTS, values[0], values[2])[1])[0], NAMES))
                window['-TABLE-'].update(num_rows=(len(calculate_grades(values[0], values[2], POINTS, float(values[3]), print_sufficient_percentage(POINTS, values[0], values[2])[1])[0])))

                if normway == "percentage":

                    image1 = os.path.join(
                        values["-FOLDER-"], save_figure(fig1, fig2,
                                group, values[4], normway, values[0], values["-FOLDER-"])[0])

                    image2 = os.path.join(
                        values["-FOLDER-"], save_figure(fig1, fig2,
                                group, values[4], normway, values[0], values["-FOLDER-"])[1])

                    window['-IMAGE1-'].update(data=convert_to_bytes(image1, resize=(window_width / 3, window_height / 2)))

                    window['-IMAGE2-'].update(data=convert_to_bytes(image2, resize=(window_width / 3, window_height / 2)))

                    os.remove(image1)
                    os.remove(image2)

                elif normway == "punten":

                    if values[0] == "":
                        sg.popup("values0")

                    image1 = os.path.join(
                        values["-FOLDER-"], save_figure(fig1, fig2,
                                group, values[4], normway, values[1], values["-FOLDER-"])[0])

                    image2 = os.path.join(
                        values["-FOLDER-"], save_figure(fig1, fig2,
                                group, values[4], normway, values[1], values["-FOLDER-"])[1])

                    window['-IMAGE1-'].update(data=convert_to_bytes(image1, resize=(window_width / 3, window_height / 2)))

                    window['-IMAGE2-'].update(data=convert_to_bytes(image2, resize=(window_width / 3, window_height / 2)))

                    os.remove(image1)
                    os.remove(image2)

                fig1_width = (fig1.get_size_inches()[0]*fig1.dpi)
                fig2_width = (fig2.get_size_inches()[0]*fig2.dpi)
                fig1_heigth = (fig1.get_size_inches()[1]*fig1.dpi)
                fig2_heigth = (fig2.get_size_inches()[1]*fig2.dpi)

                set_column_size(window['-LEFTCOLUMN-'], (window_width / 3, window_height))
                set_column_size(window['-CENTERCOLUMN-'], (window_width / 3.5, window_height))
                set_column_size(window['-RIGHTCOLUMN-'], (window_width / 3 , window_height))

                if window_width < 1180:
                    set_column_size(window['-LEFTCOLUMN-'], (350, window_height))
                    set_column_size(window['-CENTERCOLUMN-'], (350, window_height))
                    set_column_size(window['-RIGHTCOLUMN-'], (window_width - 700, window_height))


                #fig1 = draw_figure(image_element, draw_normLines(values[0], values[2], POINTS, float(values[3]),
                 #                                                       print_sufficient_percentage(POINTS, values[0],
                  #                                                                                  values[2])[1])[0])

                #fig2 = draw_figure(image_element2, draw_normLines(values[0], values[2], POINTS, float(values[2]),
                 #                                                       print_sufficient_percentage(POINTS, values[0],
                  #                                                                                values[2])[1])[1])



                #image_element.set_size(resize_image(fig1, window.size, fig1.get_width_height()))





            except:
                sg.popup("Er is iets fout gegaan, kijk of u alles goed heeft ingevuld")
                pass

        if event == "-GRADEBUTTON-":
            try:
                groupfilelist = []
                

                for i in fnames:
                    file = os.path.join(
                        values["-FOLDER-"], i
                    )
                    groupfilelist.append(file)
                print("groupfilelist: ", groupfilelist)

                for j in groupfilelist:

                    print("j", j)
                    wb = load_workbook(j)
                    print("tot hier=====2=======")
                    data = pd.ExcelFile(j)
                    print("tot hier=====3=======")

                    print("data", data)

                    sheet1 = pd.read_excel(data, 'T01')
                    print("sheet1")
                    sheet2 = pd.read_excel(data, 'T02')
                    print("sheet2")
                    sheet3 = pd.read_excel(data, 'T03')
                    print("sheet3")
                    sheet4 = pd.read_excel(data, 'T04')
                    print("sheet4")
                    sheet5 = pd.read_excel(data, 'HT01')
                    print("sheet5")
                    sheet6 = pd.read_excel(data, 'HT02')
                    print("sheet6")
                    sheet7 = pd.read_excel(data, 'HT03')
                    print("sheet7")
                    sheet8 = pd.read_excel(data, 'HT04')

                    print("tot hier=====4=======")


                    if values[4] == "T01":
                        POINTS = sheet1['Punten'].tolist()
                    if values[4] == "T02":
                        POINTS = sheet2['Punten'].tolist()
                    if values[4] == "T03":
                        POINTS = sheet3['Punten'].tolist()
                    if values[4] == "T04":
                        POINTS = sheet4['Punten'].tolist()

                    if values[4] == "HT01":
                        POINTS = sheet5['Punten'].tolist()
                    if values[4] == "HT02":
                        POINTS = sheet6['Punten'].tolist()
                    if values[4] == "HT03":
                        POINTS = sheet7['Punten'].tolist()
                    if values[4] == "HT04":
                        POINTS = sheet8['Punten'].tolist()


                    if values['RADIOB2'] == True:
                        values[0] = (float(values[1]) / float(values[2]) * 100)
                        normway = "punten"
                    else:
                        normway = "percentage"

                    print("points: ", POINTS)

                    grades = calculate_grades(values[0], values[2], POINTS, float(values[3]), print_sufficient_percentage(POINTS, values[0], values[2])[1])[0]
                    grades.insert(0, "cijfers")
                    print(grades)
                    add_column(wb, values[4], grades)
                    wb.save(j)

                    
                    
                sg.popup("De cijfers staan in het excel bestand")

            except:
                groupfilelist = []

                for i in fnames:
                    file = os.path.join(
                        values["-FOLDER-"], i
                    )
                    groupfilelist.append(file)


                sg.popup("er is iets fout gegaan, probeer opnieuw")

                pass

        if event == "-SAVEPIC-":
            try:
                if normway == "percentage":

                    image1 = os.path.join(
                        values["-FOLDER-"], save_figure(fig1, fig2,
                                group, values[4], normway, values[0], values["-FOLDER-"])[0])

                    image2 = os.path.join(
                        values["-FOLDER-"], save_figure(fig1, fig2,
                                group, values[4], normway, values[0], values["-FOLDER-"])[1])


                elif normway == "punten":

                    image1 = os.path.join(
                        values["-FOLDER-"], save_figure(fig1, fig2,
                                group, values[4], normway, values[1], values["-FOLDER-"])[0])

                    image2 = os.path.join(
                        values["-FOLDER-"], save_figure(fig1, fig2,
                                group, values[4], normway, values[1], values["-FOLDER-"])[1])

                sg.popup("De afbeeldingen zijn opgeslagen")

            except:

                sg.popup("er is iets fout gegaan, probeer opnieuw")
                pass

    window.close()

runApp()